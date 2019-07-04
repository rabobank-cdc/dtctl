# pylint: disable=W0212
"""Common functions for reporting requirements"""
import openpyxl
import openpyxl.utils
from openpyxl import styles
from openpyxl import load_workbook
from openpyxl.worksheet import table


TABLE_NAME = 'RawDataTable'


def format_report(breaches_df, output_file, template, output_format):
    """
    Format Pandas dataframe and create an excel file that holds all model breaches for which
    comments have been entered.

    :param breaches_df: Pandas DataFrame that contains all model breaches to report on
    :param output_file: Filename in String where the report should be saved to
    :param template: Excel template file for appending data (i.e. for pre-made pivot tables, reports, etc).
                     Note that a 'RawData' sheet must be present and have the same column names as the
                     data requested
    :param output_format: The output format. Does not work in combination with template if CSV is given
    :return: None
    """
    # If format is CSV, we use pandas CSV function to output CSV file and return
    if output_format == 'csv':
        breaches_df.to_csv(output_file)
        return None

    # If a template is provided, we add the data to the RawData sheet.
    if template:
        work_book = load_workbook(template)
        if 'RawData' not in work_book.sheetnames:
            raise SystemExit('No sheet with the name "RawData" found.')
        work_sheet = work_book.get_sheet_by_name('RawData')

        # Get maximum nr of rows currently in the sheet
        sheet_rows = work_sheet.max_row

        # First row is header row. We need cell values to get column names
        column_names = get_header_column_names(work_sheet)

        # build an array of all column names. We need to add breach_id manually
        # because the index of the df is named pbid and not excluded automatically
        df_columns = list(breaches_df) + ['breach_id']

        # check to ensure template has same columns as dataframe. Otherwise
        # appending creates weird results.
        if not set(column_names) == set(df_columns):
            raise SystemExit('Template file has different columns than requested report\n'
                             'Expected columns: {0}'.format(', '.join(df_columns)))

        table_ref = get_table_ref(work_sheet)  # Returns None if no table is found

        # table_ref can be 0, which is falsy
        if table_ref is not None:
            column_letter = openpyxl.utils.get_column_letter(len(breaches_df.columns) + 1)
            length_of_table = len(breaches_df.index) + sheet_rows
            tab = set_table(column_letter, length_of_table)
            work_sheet._tables[table_ref] = tab

    else:
        work_book = openpyxl.Workbook()
        work_sheet = work_book.active
        column_letter = openpyxl.utils.get_column_letter(len(breaches_df.columns) + 1)
        length_of_table = len(breaches_df.index) + 1
        tab = set_table(column_letter, length_of_table)
        work_sheet.add_table(tab)

        # No template, means new workbook and sheet.
        # So we create a table and this will be its header row
        work_sheet.append(['breach_id'] + list(breaches_df.columns))

    for breaches_row in breaches_df.iterrows():
        work_sheet.append([str(breaches_row[0])] + list(breaches_row[1]))

    set_auto_size_columns(work_sheet)

    for column_nr in range(1, len(breaches_df.columns) + 2):
        for row_nr in range(1, len(breaches_df.index) + 2):
            cell = work_sheet[f'{openpyxl.utils.get_column_letter(column_nr)}{row_nr}']
            cell.alignment = styles.Alignment(wrap_text=True)

    work_book.save(output_file)
    return None


def set_auto_size_columns(work_sheet):
    """
    Configure width for all columns in a worksheet. Workaround for the lack of
    a proper auto_width function

    :param work_sheet: Worksheet for which to configure column width
    :return: None
    """
    for col in work_sheet.columns:
        max_length = 0
        column = col[0].column  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = (max_length + 4) * 1.2
        work_sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = adjusted_width


def set_table(column_letter, length_of_table):
    """
    Format Pandas dataframe and create an excel file that holds all model breaches for which
    comments have been entered.

    :param column_letter: the final column letter for data range of table
    :param length_of_table: the final row for data range of table
    :return: Table: Returns an openpyxl Excel table with style and ranges configured.
    """
    # Set generic table style
    style = table.TableStyleInfo(name="TableStyleLight8", showFirstColumn=False,
                                 showLastColumn=False, showRowStripes=True, showColumnStripes=True)

    tab = table.Table(displayName=TABLE_NAME,
                      ref=f"A1:{column_letter}{length_of_table}")

    tab.tableStyleInfo = style
    return tab


def get_table_ref(work_sheet):
    """
    Searches all tables in a worksheet and returns the table reference for TABLE_NAME

    :param work_sheet: Worksheet that potentially contains a table
    :return: table_ref: Table reference to table of name TABLE_NAME. Returns None if not found
    """
    table_ref = None
    for i, datatable in enumerate(work_sheet._tables):
        if datatable.name == TABLE_NAME:
            table_ref = i
    return table_ref


def get_header_column_names(work_sheet):
    """
    Returns a list that contains all column names.

    :param work_sheet: Worksheet for which to return column header names.
    :return: list: List of column header names
    """
    column_names = []
    for cell in work_sheet[1]:
        column_names.append(cell.value)
    return column_names


def device_info(components, info):
    """
    Return comma separated string of values in dict

    :param components: Components to sort through
    :type components: Dict
    :param info:
    :type info: String
    :return: Comma separated string of values selected with key
    :rtype: String
    """
    return ', '. join({str(comp['device'].get(info)) for comp in components})
